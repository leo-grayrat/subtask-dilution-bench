import json
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

import openpyxl

from benchmarks.context_integration.s09_reversed_deal.materialize import materialize


DEAL = "4505"


def _make_deal_log(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deal Log"
    ws.append(["Deal#", "Customer Last", "Funding Status", "Fund Date", "Title Status", "Salesperson", "Front Gross", "Notes"])
    ws.append([4505, "Park", "Rejected", None, "Incomplete", "Carlos Vega", 4200, "Lender rejected 04/14 — unwind in progress."])
    ws.append([4506, "Mitchell", "Funded", datetime(2026, 4, 12), "Complete", "Megan Price", 1750, "keep"])
    wb.save(path)


def _make_source(root: Path) -> None:
    workspace = root / "environment/initial_workspace"
    mail_dir = root / "environment/initial_external_services/google_mail"
    slack_dir = root / "environment/initial_external_services/slack"
    tests = root / "tests"
    workspace.mkdir(parents=True)
    mail_dir.mkdir(parents=True)
    slack_dir.mkdir(parents=True)
    tests.mkdir(parents=True)
    (root / "instruction.md").write_text("Process all commission sheets.\n", encoding="utf-8")
    _make_deal_log(workspace / "Deal_Log_April2026.xlsx")
    (workspace / "CommSheet_NewVehicles_Apr1-15_OliviaMercer.xlsx").write_bytes(b"commission-sheet")
    (workspace / "keep.txt").write_text("keep", encoding="utf-8")
    inbox = {
        "emails": [
            {
                "email_id": "16",
                "subject": "Deal 4505 Unwind Complete - David Park / Ally Financial",
                "from_addr": "jasmine.patel@sunshineandsetauto.com",
                "body_text": "Deal 4505 unwind complete. Deal posting reversed in deal log. Please note for commission processing.",
                "is_read": False,
            },
            {"email_id": "keep", "subject": "Other", "body_text": "keep"},
        ]
    }
    (mail_dir / "inbox.json").write_text(json.dumps(inbox), encoding="utf-8")
    slack = {
        "channels": {"C1": {"name": "deal-desk"}},
        "messages": {
            "C1": [
                {"ts": "1", "text": "Deal 4505 unwind is happening. Ally Financial rejected funding."},
                {"ts": "2", "text": "keep"},
            ]
        },
    }
    (slack_dir / "slack.json").write_text(json.dumps(slack), encoding="utf-8")
    rubrics = [
        {"id": "target-old", "rubric_text": "Deal 4505 must be removed", "verifier_code": "OLD"},
        {"id": "other", "rubric_text": "Deal 4492 chargeback", "verifier_code": "OTHER"},
    ]
    (tests / "rubrics.json").write_text(json.dumps(rubrics), encoding="utf-8")


def _deal_row(root: Path) -> list:
    path = root / "environment/initial_workspace/Deal_Log_April2026.xlsx"
    ws = openpyxl.load_workbook(path, data_only=False)["Deal Log"]
    return [cell.value for cell in ws[2]]


def _source_message(root: Path) -> tuple[dict, dict]:
    inbox = json.loads((root / "environment/initial_external_services/google_mail/inbox.json").read_text())
    slack = json.loads((root / "environment/initial_external_services/slack/slack.json").read_text())
    return inbox["emails"][0], slack["messages"]["C1"][0]


def _run_verifier(code: str, external: Path) -> dict:
    namespace = {}
    exec(code, namespace)
    return namespace["verify"](external.parent / "workspace", external)


def _result(root: Path, variant: str) -> Path:
    root.mkdir(parents=True)
    if variant == "A":
        payroll = "April 1-15 commissions: Deal 4505 excluded because the deal was unwound."
        messages = [{"text": "Olivia Mercer: remove Deal 4505; it was unwound and cannot be included."}]
    else:
        payroll = "April 1-15 commissions: Deal 4505 — Carlos Vega — $840."
        messages = []
    (root / "mailbox.json").write_text(json.dumps({
        "emails": [{
            "folder": "Sent",
            "to_addr": "hannah.price@sunshineandsetauto.com",
            "subject": "April 1-15 payroll submission",
            "body_text": payroll,
        }]
    }), encoding="utf-8")
    (root / "slack_data.json").write_text(json.dumps({
        "channels": {"C1": {"name": "acct-payroll-commissions"}},
        "messages": {"C1": messages},
    }), encoding="utf-8")
    return root


class MaterializeS09Tests(unittest.TestCase):
    def test_local_a_keeps_reversed_state_and_only_target_rubrics(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src, out = base / "src", base / "out"
            _make_source(src)

            materialize(src, out, mode="local", variant="A")

            self.assertEqual(_deal_row(out)[2], "Rejected")
            self.assertIn("Unwind", _source_message(out)[0]["subject"])
            self.assertIn(DEAL, (out / "instruction.md").read_text(encoding="utf-8"))
            rubrics = json.loads((out / "tests/rubrics.json").read_text())
            self.assertEqual({r["id"] for r in rubrics}, {"S09-PAYROLL", "S09-NOTIFICATION"})

    def test_variant_b_changes_only_the_target_deal_state_across_its_records(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src, out = base / "src", base / "out"
            _make_source(src)

            materialize(src, out, mode="full", variant="B")

            before, after = _deal_row(src), _deal_row(out)
            self.assertEqual(after[:2], before[:2])
            self.assertEqual(after[2], "Funded")
            self.assertEqual(after[3], datetime(2026, 4, 14))
            self.assertEqual(after[4], "Complete")
            self.assertEqual(after[5:7], before[5:7])
            self.assertIn("funded", after[7].lower())
            email, slack = _source_message(out)
            self.assertIn("Funding Complete", email["subject"])
            self.assertNotIn("unwind", email["body_text"].lower())
            self.assertIn("funded", slack["text"].lower())
            self.assertEqual((out / "environment/initial_workspace/keep.txt").read_text(), "keep")

    def test_full_mode_keeps_unrelated_rubric_and_replaces_target_scoring(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src, out = base / "src", base / "out"
            _make_source(src)

            materialize(src, out, mode="full", variant="B")

            rubrics = {r["id"]: r for r in json.loads((out / "tests/rubrics.json").read_text())}
            self.assertIn("other", rubrics)
            self.assertNotIn("target-old", rubrics)
            self.assertIn("S09-PAYROLL", rubrics)
            self.assertIn("S09-NOTIFICATION", rubrics)

    def test_inconsistent_source_state_fails_loudly(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src = base / "src"
            _make_source(src)
            inbox_path = src / "environment/initial_external_services/google_mail/inbox.json"
            inbox = json.loads(inbox_path.read_text())
            inbox["emails"][0]["subject"] = "Deal 4505 Funding Complete"
            inbox_path.write_text(json.dumps(inbox), encoding="utf-8")
            with self.assertRaises(RuntimeError):
                materialize(src, base / "out", mode="full", variant="B")

    def test_embedded_verifiers_accept_correct_a_and_b_results(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src = base / "src"
            _make_source(src)
            for variant in ("A", "B"):
                task = base / f"task_{variant}"
                materialize(src, task, mode="local", variant=variant)
                rubrics = {r["id"]: r for r in json.loads((task / "tests/rubrics.json").read_text())}
                external = _result(base / f"result_{variant}", variant)
                self.assertTrue(_run_verifier(rubrics["S09-PAYROLL"]["verifier_code"], external)["pass"])
                self.assertTrue(_run_verifier(rubrics["S09-NOTIFICATION"]["verifier_code"], external)["pass"])

    def test_payroll_verifier_rejects_crossed_result(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src = base / "src"
            _make_source(src)
            task = base / "task"
            materialize(src, task, mode="local", variant="A")
            rubric = {r["id"]: r for r in json.loads((task / "tests/rubrics.json").read_text())}["S09-PAYROLL"]
            self.assertFalse(_run_verifier(rubric["verifier_code"], _result(base / "result", "B"))["pass"])

    def test_invalid_mode_or_variant_is_rejected(self):
        with tempfile.TemporaryDirectory() as td:
            src = Path(td) / "src"
            _make_source(src)
            with self.assertRaises(ValueError):
                materialize(src, Path(td) / "x", mode="tiny", variant="A")
            with self.assertRaises(ValueError):
                materialize(src, Path(td) / "y", mode="full", variant="C")


if __name__ == "__main__":
    unittest.main()
