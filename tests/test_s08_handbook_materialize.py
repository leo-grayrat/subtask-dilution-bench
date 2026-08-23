import json
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

import openpyxl

from benchmarks.context_integration.s08_expired_agreement.materialize import materialize


INVOICE_NUMBER = "CC-2026-0168"
VENDOR_ID = "V-000031"


def _save_workbook(path: Path, sheets: dict[str, list[list]]) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    wb.save(path)


def _make_source(root: Path) -> None:
    workspace = root / "environment/initial_workspace"
    tests = root / "tests"
    workspace.mkdir(parents=True)
    tests.mkdir(parents=True)
    (root / "instruction.md").write_text("Process every pending invoice.\n", encoding="utf-8")

    _save_workbook(workspace / "recurring_agreements.xlsx", {
        "Recurring Agreements": [
            ["Agreement ID", "Vendor ID", "Vendor Name", "Expected Amount ($)", "Agreement Expiration Date", "Status", "Notes"],
            ["RA-2026-001", VENDOR_ID, "Cascade Cleaning Co", 2200, datetime(2026, 3, 31), "Expired", "EXPIRED 03/31/2026"],
            ["RA-KEEP", "V-OTHER", "Other Vendor", 100, datetime(2027, 1, 1), "Active", "keep"],
        ]
    })
    _save_workbook(workspace / "ap_ledger.xlsx", {
        "Invoice Register": [
            ["Invoice ID", "Vendor ID", "Invoice Number", "Invoice Amt ($)", "Status", "Hold Code", "Credit Applied ($)"],
            ["INV-026", "V-OTHER", "OTHER-1", 100, "PENDING", None, 0],
        ],
        "Payment Queue": [
            ["Payment ID", "Invoice ID", "Vendor ID", "Invoice Number", "Invoice Amt ($)", "Payment Amt ($)"],
        ],
    })
    _save_workbook(workspace / "vendor_master.xlsx", {
        "Vendor Master": [
            ["Vendor ID", "Vendor Name", "Status", "Outstanding Credit"],
            [VENDOR_ID, "Cascade Cleaning Co", "Active", 150],
        ]
    })
    (workspace / "INV-CC-2026-0168.pdf").write_bytes(b"invoice")
    (workspace / "keep.txt").write_text("keep", encoding="utf-8")
    rubrics = [
        {"id": "old-target", "rubric_text": f"{INVOICE_NUMBER} must be held", "verifier_code": "OLD"},
        {"id": "other", "rubric_text": "INV-018 must be approved", "verifier_code": "OTHER"},
    ]
    (tests / "rubrics.json").write_text(json.dumps(rubrics), encoding="utf-8")


def _agreement_row(path: Path) -> list:
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Recurring Agreements"]
    return [cell.value for cell in ws[2]]


def _run_verifier(code: str, workspace: Path, external: Path) -> dict:
    namespace = {}
    exec(code, namespace)
    return namespace["verify"](workspace, external)


def _make_result(root: Path, variant: str) -> tuple[Path, Path]:
    workspace = root / "workspace"
    external = root / "external"
    workspace.mkdir(parents=True)
    external.mkdir(parents=True)
    if variant == "A":
        invoice_row = ["INV-027", VENDOR_ID, INVOICE_NUMBER, 2200, "ON HOLD", "APPR", 0]
        queue_rows = [["Payment ID", "Invoice ID", "Vendor ID", "Invoice Number", "Invoice Amt ($)", "Payment Amt ($)"]]
        credit = 150
        emails = [{
            "folder": "sent",
            "to_addr": "t.okonkwo@meridianpartners.com",
            "subject": f"Meridian Partners - Expired Recurring Agreement - Cascade Cleaning - Invoice {INVOICE_NUMBER}",
            "body_text": "Agreement expired March 31 2026. Invoice placed on APPR hold pending renewal.",
        }]
        slack_messages = []
    else:
        invoice_row = ["INV-027", VENDOR_ID, INVOICE_NUMBER, 2200, "APPROVED", None, 150]
        queue_rows = [
            ["Payment ID", "Invoice ID", "Vendor ID", "Invoice Number", "Invoice Amt ($)", "Payment Amt ($)"],
            ["PMT-027", "INV-027", VENDOR_ID, INVOICE_NUMBER, 2200, 2050],
        ]
        credit = 0
        emails = []
        slack_messages = [{"text": f"SELF-APPROVED: Cascade Cleaning, {VENDOR_ID}, Invoice {INVOICE_NUMBER}"}]

    _save_workbook(workspace / "ap_ledger.xlsx", {
        "Invoice Register": [
            ["Invoice ID", "Vendor ID", "Invoice Number", "Invoice Amt ($)", "Status", "Hold Code", "Credit Applied ($)"],
            invoice_row,
        ],
        "Payment Queue": queue_rows,
    })
    _save_workbook(workspace / "vendor_master.xlsx", {
        "Vendor Master": [
            ["Vendor ID", "Vendor Name", "Status", "Outstanding Credit"],
            [VENDOR_ID, "Cascade Cleaning Co", "Active", credit],
        ]
    })
    (external / "mailbox.json").write_text(json.dumps({"emails": emails}), encoding="utf-8")
    (external / "slack_data.json").write_text(json.dumps({
        "channels": {"C1": {"name": "ap-approvals"}},
        "messages": {"C1": slack_messages},
    }), encoding="utf-8")
    return workspace, external


class MaterializeS08Tests(unittest.TestCase):
    def test_local_a_keeps_expired_agreement_and_only_target_rubrics(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src, out = base / "src", base / "out"
            _make_source(src)

            materialize(src, out, mode="local", variant="A")

            row = _agreement_row(out / "environment/initial_workspace/recurring_agreements.xlsx")
            self.assertEqual(row[4], datetime(2026, 3, 31))
            self.assertEqual(row[5], "Expired")
            self.assertIn(INVOICE_NUMBER, (out / "instruction.md").read_text(encoding="utf-8"))
            rubrics = json.loads((out / "tests/rubrics.json").read_text(encoding="utf-8"))
            self.assertEqual({r["id"] for r in rubrics}, {"S08-INVOICE", "S08-COMMUNICATIONS"})

    def test_variant_b_only_renews_the_target_agreement_row(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src, out = base / "src", base / "out"
            _make_source(src)

            materialize(src, out, mode="full", variant="B")

            before = _agreement_row(src / "environment/initial_workspace/recurring_agreements.xlsx")
            after = _agreement_row(out / "environment/initial_workspace/recurring_agreements.xlsx")
            self.assertEqual(after[:4], before[:4])
            self.assertEqual(after[4], datetime(2027, 3, 31))
            self.assertEqual(after[5], "Active")
            self.assertIn("2027", after[6])
            wb = openpyxl.load_workbook(out / "environment/initial_workspace/recurring_agreements.xlsx")
            self.assertEqual(wb["Recurring Agreements"]["A3"].value, "RA-KEEP")
            self.assertEqual((out / "environment/initial_workspace/keep.txt").read_text(), "keep")

    def test_full_mode_keeps_unrelated_rubric_and_replaces_target_scoring(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src, out = base / "src", base / "out"
            _make_source(src)

            materialize(src, out, mode="full", variant="B")

            rubrics = {r["id"]: r for r in json.loads((out / "tests/rubrics.json").read_text())}
            self.assertIn("other", rubrics)
            self.assertNotIn("old-target", rubrics)
            self.assertIn("S08-INVOICE", rubrics)
            self.assertIn("S08-COMMUNICATIONS", rubrics)

    def test_wrong_source_agreement_state_fails_loudly(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src = base / "src"
            _make_source(src)
            path = src / "environment/initial_workspace/recurring_agreements.xlsx"
            wb = openpyxl.load_workbook(path)
            wb["Recurring Agreements"]["F2"] = "Active"
            wb.save(path)

            with self.assertRaises(RuntimeError):
                materialize(src, base / "out", mode="full", variant="B")

    def test_embedded_verifiers_accept_correct_a_and_b_results(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src = base / "src"
            _make_source(src)
            for variant in ("A", "B"):
                task = base / f"task_{variant}"
                materialize(src, task, mode="local", variant=variant)
                rubrics = {r["id"]: r for r in json.loads((task / "tests/rubrics.json").read_text())}
                workspace, external = _make_result(base / f"result_{variant}", variant)
                self.assertTrue(_run_verifier(rubrics["S08-INVOICE"]["verifier_code"], workspace, external)["pass"])
                self.assertTrue(_run_verifier(rubrics["S08-COMMUNICATIONS"]["verifier_code"], workspace, external)["pass"])

    def test_embedded_invoice_verifier_rejects_crossed_result(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src = base / "src"
            _make_source(src)
            task = base / "task"
            materialize(src, task, mode="local", variant="A")
            rubric = {r["id"]: r for r in json.loads((task / "tests/rubrics.json").read_text())}["S08-INVOICE"]
            workspace, external = _make_result(base / "result", "B")
            self.assertFalse(_run_verifier(rubric["verifier_code"], workspace, external)["pass"])

    def test_invalid_mode_or_variant_is_rejected(self):
        with tempfile.TemporaryDirectory() as td:
            src = Path(td) / "src"
            _make_source(src)
            with self.assertRaises(ValueError):
                materialize(src, Path(td) / "x", mode="tiny", variant="A")
            with self.assertRaises(ValueError):
                materialize(src, Path(td) / "y", mode="full", variant="C")


if __name__ == "__main__":
    unittest.main()
