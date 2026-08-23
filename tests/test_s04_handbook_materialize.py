import io
import json
import tempfile
import unittest
import zipfile
from pathlib import Path

import openpyxl

from benchmarks.context_integration.s04_contact_history.materialize import materialize


CASE = "Rodriguez_07221975"


def _intake_bytes(log: str, consent: str = "Pending") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["CASE_ID", "Patient_Email", "Physician_Email", "Contact_Attempt_Log", "Consent_Status"])
    ws.append([CASE, "c.rodriguez75@webmail.net", "rcho@greenmedclinic.org", log, consent])
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _make_source(root: Path) -> None:
    workspace = root / "environment/initial_workspace"
    workspace.mkdir(parents=True)
    (root / "tests").mkdir(parents=True)
    (root / "instruction.md").write_text("Process Webb and the other patient.\n", encoding="utf-8")
    log = (
        "2026-03-25 10:15 — Voicemail\n"
        "2026-03-26 14:30 — No Answer\n"
        "2026-03-28 09:00 — Voicemail"
    )
    with zipfile.ZipFile(workspace / f"{CASE}.zip", "w") as zf:
        zf.writestr(f"{CASE}/intake.xlsx", _intake_bytes(log))
        zf.writestr(f"{CASE}/audit_log.xlsx", b"keep-audit")
    (workspace / "keep.txt").write_text("keep", encoding="utf-8")
    rubrics = [
        {"id": "webb", "rubric_text": "Webb_03151968 output", "verifier_code": "WEBB"},
        {"id": "rodriguez", "rubric_text": "Rodriguez_07221975 must be unreachable", "verifier_code": "OLD"},
    ]
    (root / "tests/rubrics.json").write_text(json.dumps(rubrics), encoding="utf-8")


def _read_intake(zip_path: Path):
    with zipfile.ZipFile(zip_path) as zf:
        data = zf.read(f"{CASE}/intake.xlsx")
        audit = zf.read(f"{CASE}/audit_log.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    ws = wb.active
    headers = {cell.value: i for i, cell in enumerate(ws[1])}
    return ws.cell(2, headers["Contact_Attempt_Log"] + 1).value, audit


def _run_verifier(code: str, workspace: Path, external: Path) -> dict:
    namespace = {}
    exec(code, namespace)
    return namespace["verify"](workspace, external)


class MaterializeS04Tests(unittest.TestCase):
    def test_local_a_keeps_materials_and_uses_only_target_rubrics(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src, out = base / "src", base / "out"
            _make_source(src)

            materialize(src, out, mode="local", variant="A")

            self.assertEqual((out / "environment/initial_workspace/keep.txt").read_text(), "keep")
            self.assertIn(CASE, (out / "instruction.md").read_text())
            rubrics = json.loads((out / "tests/rubrics.json").read_text())
            self.assertEqual({r["id"] for r in rubrics}, {"S04-CASE-STATE", "S04-COMMUNICATIONS"})

    def test_variant_b_changes_only_one_contact_result_inside_zip(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src, out = base / "src", base / "out"
            _make_source(src)

            materialize(src, out, mode="full", variant="B")

            original_log, original_audit = _read_intake(src / f"environment/initial_workspace/{CASE}.zip")
            changed_log, changed_audit = _read_intake(out / f"environment/initial_workspace/{CASE}.zip")
            self.assertEqual(changed_log.count("\n"), 2)
            self.assertEqual(original_log.splitlines()[:2], changed_log.splitlines()[:2])
            self.assertTrue(original_log.splitlines()[2].endswith("Voicemail"))
            self.assertTrue(changed_log.splitlines()[2].endswith("Reached"))
            self.assertEqual(original_audit, changed_audit)

    def test_full_mode_keeps_unrelated_rubrics_and_replaces_target_scoring(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src, out = base / "src", base / "out"
            _make_source(src)

            materialize(src, out, mode="full", variant="B")

            rubrics = {r["id"]: r for r in json.loads((out / "tests/rubrics.json").read_text())}
            self.assertIn("webb", rubrics)
            self.assertNotIn("rodriguez", rubrics)
            self.assertIn("S04-CASE-STATE", rubrics)
            self.assertIn("S04-COMMUNICATIONS", rubrics)

    def test_missing_three_attempt_history_fails_loudly(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src = base / "src"
            _make_source(src)
            zip_path = src / f"environment/initial_workspace/{CASE}.zip"
            with zipfile.ZipFile(zip_path, "w") as zf:
                zf.writestr(f"{CASE}/intake.xlsx", _intake_bytes("2026-03-25 — Voicemail"))
                zf.writestr(f"{CASE}/audit_log.xlsx", b"keep-audit")
            with self.assertRaises(RuntimeError):
                materialize(src, base / "out", mode="full", variant="B")

    def test_embedded_verifiers_accept_correct_a_and_b_results(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src = base / "src"
            _make_source(src)
            for variant in ("A", "B"):
                task = base / f"task_{variant}"
                materialize(src, task, mode="local", variant=variant)
                rubrics = {r["id"]: r for r in json.loads((task / "tests/rubrics.json").read_text())}
                workspace = base / f"workspace_{variant}" / CASE
                external = base / f"external_{variant}"
                workspace.mkdir(parents=True)
                external.mkdir()
                source_log, _ = _read_intake(task / f"environment/initial_workspace/{CASE}.zip")
                (workspace / "intake.xlsx").write_bytes(
                    _intake_bytes(source_log, "Pending" if variant == "A" else "Form Sent")
                )
                if variant == "A":
                    slack_messages = [{"text": f"[UNREACHABLE] {CASE} — 3 failed attempts"}]
                    emails = [{
                        "folder": "sent",
                        "to_addr": "rcho@greenmedclinic.org",
                        "subject": f"Unable to Contact Patient | CASE_ID: {CASE}",
                    }]
                else:
                    slack_messages = []
                    emails = [{
                        "folder": "sent",
                        "to_addr": "c.rodriguez75@webmail.net",
                        "subject": "CareIG — Consent Form Required | Rodriguez",
                    }]
                (external / "slack_data.json").write_text(json.dumps({
                    "channels": {"C1": {"name": "intake-alerts"}},
                    "messages": {"C1": slack_messages},
                }), encoding="utf-8")
                (external / "mailbox.json").write_text(json.dumps({"emails": emails}), encoding="utf-8")
                self.assertTrue(_run_verifier(rubrics["S04-CASE-STATE"]["verifier_code"], workspace.parent, external)["pass"])
                self.assertTrue(_run_verifier(rubrics["S04-COMMUNICATIONS"]["verifier_code"], workspace.parent, external)["pass"])

    def test_invalid_mode_or_variant_is_rejected(self):
        with tempfile.TemporaryDirectory() as td:
            src = Path(td) / "src"
            _make_source(src)
            with self.assertRaises(ValueError):
                materialize(src, Path(td) / "x", mode="tiny", variant="A")
            with self.assertRaises(ValueError):
                materialize(src, Path(td) / "y", mode="full", variant="C")


if __name__ == "__main__":
    unittest.main()
