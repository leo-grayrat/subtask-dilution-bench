from __future__ import annotations

import io
import json
import shutil
import zipfile
from pathlib import Path

import openpyxl


CASE_ID = "Rodriguez_07221975"
ZIP_NAME = f"{CASE_ID}.zip"
LOCAL_INSTRUCTION = (
    "Process only the current intake state for patient case Rodriguez_07221975. "
    "Read the existing case files and CareIG SOP, then complete every action required for this case. "
    "Use Jake Morrow as the staff name where a signature is required.\n"
)


def _state_verifier(variant: str) -> str:
    return f'''from pathlib import Path
import openpyxl

def verify(workspace_path, external_services_path=None):
    root = Path(workspace_path)
    folder = root / "Rodriguez_07221975"
    if not folder.is_dir():
        matches = [p for p in root.rglob("Rodriguez_07221975") if p.is_dir()]
        if not matches:
            return {{"pass": False, "score": 0.0, "feedback": "Rodriguez case folder not found"}}
        folder = matches[0]
    intake = folder / "intake.xlsx"
    if not intake.exists():
        return {{"pass": False, "score": 0.0, "feedback": "intake.xlsx not found"}}
    wb = openpyxl.load_workbook(intake, data_only=False)
    ws = wb.active
    headers = {{str(c.value or "").strip().lower(): i for i, c in enumerate(ws[1])}}
    log_col = headers.get("contact_attempt_log")
    consent_col = headers.get("consent_status")
    if log_col is None or consent_col is None:
        return {{"pass": False, "score": 0.0, "feedback": "required intake columns missing"}}
    row = None
    for values in ws.iter_rows(min_row=2, values_only=True):
        if values and "rodriguez_07221975" in str(values[0] or "").lower():
            row = values
            break
    if row is None:
        return {{"pass": False, "score": 0.0, "feedback": "Rodriguez intake row missing"}}
    log = str(row[log_col] or "")
    consent = str(row[consent_col] or "").strip().lower()
    lines = [x for x in log.splitlines() if x.strip()]
    failures = []
    if len(lines) != 3:
        failures.append(f"expected 3 contact entries, found {{len(lines)}}")
    if "{variant}" == "A":
        if any("reached" in x.lower() and "not reached" not in x.lower() for x in lines):
            failures.append("variant A incorrectly contains a reached attempt")
        if consent == "form sent":
            failures.append("variant A incorrectly marked consent form sent")
    else:
        if not any(x.strip().lower().endswith("reached") for x in lines):
            failures.append("variant B lacks a reached attempt")
        if consent != "form sent":
            failures.append(f"Consent_Status={{consent!r}}, expected Form Sent")
    return {{
        "pass": not failures,
        "score": 1.0 if not failures else 0.0,
        "feedback": "ok" if not failures else "; ".join(failures),
    }}
'''


def _communications_verifier(variant: str) -> str:
    return f'''from pathlib import Path
import json

def verify(workspace_path, external_services_path=None):
    if external_services_path is None:
        return {{"pass": False, "score": 0.0, "feedback": "external services path missing"}}
    root = Path(external_services_path)
    slack_path = root / "slack_data.json"
    mailbox_path = root / "mailbox.json"
    if not slack_path.exists() or not mailbox_path.exists():
        return {{"pass": False, "score": 0.0, "feedback": "Slack or mailbox result missing"}}
    slack = json.loads(slack_path.read_text(encoding="utf-8"))
    channel_ids = [
        cid for cid, info in slack.get("channels", {{}}).items()
        if isinstance(info, dict) and str(info.get("name", "")).lower() == "intake-alerts"
    ]
    alert_text = " ".join(
        str(message.get("text", "")) if isinstance(message, dict) else str(message)
        for cid in channel_ids
        for message in slack.get("messages", {{}}).get(cid, [])
    ).lower()
    emails = json.loads(mailbox_path.read_text(encoding="utf-8")).get("emails", [])
    sent = [e for e in emails if str(e.get("folder", "")).lower() == "sent"]
    unreachable_alert = "[unreachable]" in alert_text and "rodriguez_07221975" in alert_text
    physician_notice = any(
        "rcho@greenmedclinic.org" in str(e.get("to_addr", "")).lower()
        and "unable to contact patient" in str(e.get("subject", "")).lower()
        for e in sent
    )
    consent_email = any(
        "c.rodriguez75@webmail.net" in str(e.get("to_addr", "")).lower()
        and "consent form required" in str(e.get("subject", "")).lower()
        for e in sent
    )
    failures = []
    if "{variant}" == "A":
        if not unreachable_alert:
            failures.append("unreachable Slack alert missing")
        if not physician_notice:
            failures.append("physician unable-to-contact email missing")
        if consent_email:
            failures.append("consent email should not be sent when no attempt reached")
    else:
        if unreachable_alert:
            failures.append("reached patient was incorrectly marked unreachable")
        if physician_notice:
            failures.append("reached patient incorrectly produced physician failure notice")
        if not consent_email:
            failures.append("patient consent email missing")
    return {{
        "pass": not failures,
        "score": 1.0 if not failures else 0.0,
        "feedback": "ok" if not failures else "; ".join(failures),
    }}
'''


def _target_rubrics(variant: str) -> list[dict]:
    if variant == "A":
        state = "Rodriguez has three failed contact attempts and must remain on the unreachable path."
        communications = "Rodriguez must receive the unreachable alert and physician notification, not a consent email."
    else:
        state = "One Rodriguez contact attempt is Reached, so Consent_Status must become Form Sent."
        communications = "Rodriguez must not be marked unreachable and must receive the consent-form email."
    return [
        {
            "id": "S04-CASE-STATE",
            "sort_order": 100,
            "rubric_text": state,
            "verifier_code": _state_verifier(variant),
            "criterion_type": "expected_output",
        },
        {
            "id": "S04-COMMUNICATIONS",
            "sort_order": 101,
            "rubric_text": communications,
            "verifier_code": _communications_verifier(variant),
            "criterion_type": "expected_output",
        },
    ]


def _read_case_archive(path: Path) -> tuple[list[tuple[zipfile.ZipInfo, bytes]], int, str]:
    with zipfile.ZipFile(path, "r") as source:
        entries = [(info, source.read(info.filename)) for info in source.infolist()]
    intake_indexes = [
        i for i, (info, _) in enumerate(entries)
        if info.filename.replace("\\", "/").endswith(f"{CASE_ID}/intake.xlsx")
    ]
    if len(intake_indexes) != 1:
        raise RuntimeError(f"expected one Rodriguez intake.xlsx in archive, found {len(intake_indexes)}")
    index = intake_indexes[0]
    wb = openpyxl.load_workbook(io.BytesIO(entries[index][1]), data_only=False)
    ws = wb.active
    headers = {str(c.value or "").strip().lower(): i + 1 for i, c in enumerate(ws[1])}
    log_col = headers.get("contact_attempt_log")
    case_col = headers.get("case_id")
    if log_col is None or case_col is None:
        raise RuntimeError("Rodriguez intake workbook lacks CASE_ID or Contact_Attempt_Log")
    target_row = None
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row, case_col).value or "").strip() == CASE_ID:
            target_row = row
            break
    if target_row is None:
        raise RuntimeError("Rodriguez row not found in intake workbook")
    log = str(ws.cell(target_row, log_col).value or "")
    return entries, index, log


def _rewrite_reached_attempt(path: Path) -> None:
    entries, intake_index, log = _read_case_archive(path)
    lines = [x for x in log.splitlines() if x.strip()]
    if len(lines) != 3 or any(x.strip().lower().endswith("reached") for x in lines):
        raise RuntimeError("expected exactly three failed Rodriguez contact attempts")
    if not lines[2].strip().lower().endswith("voicemail"):
        raise RuntimeError("expected third Rodriguez attempt to end in Voicemail")
    lines[2] = lines[2][: -len("Voicemail")] + "Reached"

    wb = openpyxl.load_workbook(io.BytesIO(entries[intake_index][1]), data_only=False)
    ws = wb.active
    headers = {str(c.value or "").strip().lower(): i + 1 for i, c in enumerate(ws[1])}
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row, headers["case_id"]).value or "").strip() == CASE_ID:
            ws.cell(row, headers["contact_attempt_log"]).value = "\n".join(lines)
            break
    stream = io.BytesIO()
    wb.save(stream)
    entries[intake_index] = (entries[intake_index][0], stream.getvalue())

    archive = io.BytesIO()
    with zipfile.ZipFile(archive, "w") as output:
        for info, data in entries:
            output.writestr(info, data)
    path.write_bytes(archive.getvalue())


def materialize(source_task_dir: str | Path, output_dir: str | Path, *, mode: str, variant: str) -> None:
    if mode not in {"local", "full"}:
        raise ValueError("mode must be 'local' or 'full'")
    if variant not in {"A", "B"}:
        raise ValueError("variant must be 'A' or 'B'")

    source = Path(source_task_dir)
    output = Path(output_dir)
    if not source.is_dir():
        raise FileNotFoundError(source)
    if output.exists():
        shutil.rmtree(output)
    shutil.copytree(source, output)

    archive = output / "environment/initial_workspace" / ZIP_NAME
    _, _, log = _read_case_archive(archive)
    lines = [x for x in log.splitlines() if x.strip()]
    if len(lines) != 3 or any(x.strip().lower().endswith("reached") for x in lines):
        raise RuntimeError("source must contain exactly three failed Rodriguez contact attempts")
    if variant == "B":
        _rewrite_reached_attempt(archive)

    if mode == "local":
        (output / "instruction.md").write_text(LOCAL_INSTRUCTION, encoding="utf-8")

    rubrics_path = output / "tests/rubrics.json"
    rubrics = json.loads(rubrics_path.read_text(encoding="utf-8"))
    target = _target_rubrics(variant)
    if mode == "local":
        rubrics = target
    else:
        rubrics = [
            rubric for rubric in rubrics
            if CASE_ID.lower() not in str(rubric.get("rubric_text", "")).lower()
        ]
        rubrics.extend(target)
    rubrics_path.write_text(json.dumps(rubrics, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description="Build S04 HANDBOOK local/full A/B task variants")
    parser.add_argument("source_task_dir")
    parser.add_argument("output_dir")
    parser.add_argument("--mode", choices=["local", "full"], required=True)
    parser.add_argument("--variant", choices=["A", "B"], required=True)
    args = parser.parse_args()
    materialize(args.source_task_dir, args.output_dir, mode=args.mode, variant=args.variant)
    return 0


if __name__ == "__main__":
    raise SystemExit(_main())
