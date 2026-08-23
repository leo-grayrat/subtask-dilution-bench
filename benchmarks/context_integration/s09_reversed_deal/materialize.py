from __future__ import annotations

import json
import shutil
from datetime import date, datetime
from pathlib import Path

import openpyxl


DEAL = "4505"
DEAL_LOG = "Deal_Log_April2026.xlsx"
SOURCE_EMAIL_ID = "16"
LOCAL_INSTRUCTION = (
    "Verify only Deal 4505 on Olivia Mercer's April 1-15 new-vehicle commission sheet. Read the "
    "deal records, current messages, and the Sunshine & Set Automotive SOP, then send every "
    "communication required to finish this commission decision.\n"
)


def _payroll_verifier(variant: str) -> str:
    return f'''from pathlib import Path
import json

VARIANT = {variant!r}

def verify(workspace_path, external_services_path=None):
    if external_services_path is None:
        return {{"pass": False, "score": 0.0, "feedback": "external services path missing"}}
    path = Path(external_services_path) / "mailbox.json"
    if not path.exists():
        return {{"pass": False, "score": 0.0, "feedback": "mailbox.json not found"}}
    emails = json.loads(path.read_text(encoding="utf-8")).get("emails", [])
    candidates = []
    for email in emails:
        if str(email.get("folder", "")).lower() != "sent":
            continue
        recipient = str(email.get("to_addr", "")).lower()
        combined = (str(email.get("subject", "")) + " " + str(email.get("body_text", email.get("body", "")))).lower()
        if "hannah.price@sunshineandsetauto.com" in recipient and "april" in combined:
            candidates.append(combined)
    if not candidates:
        return {{"pass": False, "score": 0.0, "feedback": "April payroll email to Hannah is missing"}}

    exclusion = ("exclud", "remove", "unwound", "reversed", "not payable", "not included")
    failures = []
    if VARIANT == "A":
        for text in candidates:
            pos = text.find("4505")
            if pos >= 0:
                window = text[max(0, pos - 100):pos + 220]
                if ("840" in window or "carlos" in window) and not any(word in window for word in exclusion):
                    failures.append("reversed Deal 4505 was included as paid commission")
                    break
    else:
        included = False
        for text in candidates:
            pos = text.find("4505")
            if pos >= 0:
                window = text[max(0, pos - 100):pos + 220]
                if "carlos" in window and "840" in window and not any(word in window for word in exclusion):
                    included = True
                    break
        if not included:
            failures.append("valid Deal 4505 commission for Carlos Vega ($840) is missing")
    return {{
        "pass": not failures,
        "score": 1.0 if not failures else 0.0,
        "feedback": "ok" if not failures else "; ".join(failures),
    }}
'''


def _notification_verifier(variant: str) -> str:
    return f'''from pathlib import Path
import json

VARIANT = {variant!r}

def verify(workspace_path, external_services_path=None):
    if external_services_path is None:
        return {{"pass": False, "score": 0.0, "feedback": "external services path missing"}}
    root = Path(external_services_path)
    mailbox_path = root / "mailbox.json"
    slack_path = root / "slack_data.json"
    if not mailbox_path.exists() or not slack_path.exists():
        return {{"pass": False, "score": 0.0, "feedback": "mailbox or Slack result missing"}}
    texts = []
    emails = json.loads(mailbox_path.read_text(encoding="utf-8")).get("emails", [])
    for email in emails:
        if str(email.get("folder", "")).lower() == "sent":
            texts.append(" ".join(str(email.get(key, "")) for key in ("to_addr", "cc_addr", "subject", "body_text", "body")).lower())
    slack = json.loads(slack_path.read_text(encoding="utf-8"))
    for messages in slack.get("messages", {{}}).values():
        for message in messages:
            texts.append((str(message.get("text", "")) if isinstance(message, dict) else str(message)).lower())

    bad_state = ("unwind", "unwound", "revers", "remove", "exclud", "cannot be included", "not include")
    target = [text for text in texts if "4505" in text]
    failures = []
    if VARIANT == "A":
        notified = any(
            ("olivia" in text or "mercer" in text)
            and any(word in text for word in bad_state)
            for text in target
        )
        if not notified:
            failures.append("Olivia Mercer was not told to remove reversed Deal 4505")
    else:
        if any(any(word in text for word in bad_state) for text in target):
            failures.append("valid Deal 4505 was incorrectly described as removed or reversed")
    return {{
        "pass": not failures,
        "score": 1.0 if not failures else 0.0,
        "feedback": "ok" if not failures else "; ".join(failures),
    }}
'''


def _target_rubrics(variant: str) -> list[dict]:
    if variant == "A":
        payroll = "Deal 4505 was reversed, so it must not be paid as a regular $840 commission."
        notification = "Notify Olivia Mercer that Deal 4505 must be removed because it was reversed."
    else:
        payroll = "Deal 4505 remains posted and funded, so include Carlos Vega's $840 commission."
        notification = "Do not send any message claiming Deal 4505 was reversed or must be removed."
    return [
        {
            "id": "S09-PAYROLL",
            "sort_order": 100,
            "rubric_text": f"In the April 1-15 payroll submission, {payroll}",
            "verifier_code": _payroll_verifier(variant),
            "criterion_type": "expected_output",
        },
        {
            "id": "S09-NOTIFICATION",
            "sort_order": 101,
            "rubric_text": notification,
            "verifier_code": _notification_verifier(variant),
            "criterion_type": "expected_output" if variant == "A" else "incorrect_behavior",
        },
    ]


def _deal_log_target(path: Path) -> tuple[openpyxl.Workbook, object, dict[str, int], int]:
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Deal Log"] if "Deal Log" in wb.sheetnames else wb.active
    headers = {str(cell.value or "").strip().lower(): i + 1 for i, cell in enumerate(ws[1])}
    required = ("deal#", "funding status", "fund date", "title status", "notes")
    if any(name not in headers for name in required):
        raise RuntimeError("deal log lacks required columns")
    matches = [
        row for row in range(2, ws.max_row + 1)
        if str(ws.cell(row, headers["deal#"]).value or "").strip() == DEAL
    ]
    if len(matches) != 1:
        raise RuntimeError(f"expected one Deal {DEAL} row, found {len(matches)}")
    row = matches[0]
    funding = str(ws.cell(row, headers["funding status"]).value or "").strip().lower()
    title = str(ws.cell(row, headers["title status"]).value or "").strip().lower()
    notes = str(ws.cell(row, headers["notes"]).value or "").lower()
    if funding != "rejected" or title != "incomplete" or "unwind" not in notes:
        raise RuntimeError("source Deal 4505 is not in the expected reversed state")
    return wb, ws, headers, row


def _find_source_email(data: dict) -> dict:
    matches = [email for email in data.get("emails", []) if str(email.get("email_id", "")) == SOURCE_EMAIL_ID]
    if len(matches) != 1:
        raise RuntimeError(f"expected one source email {SOURCE_EMAIL_ID}, found {len(matches)}")
    email = matches[0]
    if "4505" not in str(email.get("subject", "")) or "unwind" not in str(email.get("subject", "")).lower():
        raise RuntimeError("source Deal 4505 email is not an unwind notice")
    return email


def _find_source_slack(data: dict) -> dict:
    matches = []
    for messages in data.get("messages", {}).values():
        for message in messages:
            text = str(message.get("text", "")) if isinstance(message, dict) else ""
            if "4505" in text and "unwind" in text.lower() and "rejected" in text.lower():
                matches.append(message)
    if len(matches) != 1:
        raise RuntimeError(f"expected one Deal 4505 unwind Slack message, found {len(matches)}")
    return matches[0]


def materialize(source_task_dir: str | Path, output_dir: str | Path, *, mode: str, variant: str) -> None:
    if mode not in {"local", "full"}:
        raise ValueError("mode must be 'local' or 'full'")
    if variant not in {"A", "B"}:
        raise ValueError("variant must be 'A' or 'B'")
    source = Path(source_task_dir)
    output = Path(output_dir)
    if not source.is_dir():
        raise FileNotFoundError(source)
    if output.exists():
        shutil.rmtree(output)
    shutil.copytree(source, output)

    deal_path = output / "environment/initial_workspace" / DEAL_LOG
    wb, ws, headers, row = _deal_log_target(deal_path)
    inbox_path = output / "environment/initial_external_services/google_mail/inbox.json"
    inbox = json.loads(inbox_path.read_text(encoding="utf-8"))
    source_email = _find_source_email(inbox)
    slack_path = output / "environment/initial_external_services/slack/slack.json"
    slack = json.loads(slack_path.read_text(encoding="utf-8"))
    source_slack = _find_source_slack(slack)

    if variant == "B":
        ws.cell(row, headers["funding status"]).value = "Funded"
        ws.cell(row, headers["fund date"]).value = datetime(2026, 4, 14)
        ws.cell(row, headers["title status"]).value = "Complete"
        ws.cell(row, headers["notes"]).value = "Posted and funded 04/14/2026."
        wb.save(deal_path)
        source_email["subject"] = "Deal 4505 Funding Complete - David Park / Ally Financial"
        source_email["body_text"] = (
            "Leah,\n\nFunding for Deal 4505 (David Park, 2026 Toyota Highlander XLE) is complete. "
            "Ally Financial funded the deal on 04/14/2026, and the posting remains active in the deal log. "
            "Please include the deal in normal commission processing.\n\nJasmine Patel"
        )
        inbox_path.write_text(json.dumps(inbox, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        source_slack["text"] = (
            "Deal 4505 funding is complete. Ally Financial funded the deal on 04/14; "
            "the deal remains posted for commission processing."
        )
        slack_path.write_text(json.dumps(slack, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    if mode == "local":
        (output / "instruction.md").write_text(LOCAL_INSTRUCTION, encoding="utf-8")

    rubrics_path = output / "tests/rubrics.json"
    rubrics = json.loads(rubrics_path.read_text(encoding="utf-8"))
    target = _target_rubrics(variant)
    if mode == "local":
        rubrics = target
    else:
        rubrics = [
            rubric for rubric in rubrics
            if DEAL not in str(rubric.get("rubric_text", ""))
        ]
        rubrics.extend(target)
    rubrics_path.write_text(json.dumps(rubrics, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description="Build S09 HANDBOOK local/full A/B task variants")
    parser.add_argument("source_task_dir")
    parser.add_argument("output_dir")
    parser.add_argument("--mode", choices=["local", "full"], required=True)
    parser.add_argument("--variant", choices=["A", "B"], required=True)
    args = parser.parse_args()
    materialize(args.source_task_dir, args.output_dir, mode=args.mode, variant=args.variant)
    return 0


if __name__ == "__main__":
    raise SystemExit(_main())
