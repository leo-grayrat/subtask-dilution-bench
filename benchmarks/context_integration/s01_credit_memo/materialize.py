from __future__ import annotations

import base64
import json
import re
import shutil
import zlib
from pathlib import Path

TARGET_EMAIL_ID = "63"
TARGET_EMAIL_RUBRIC_ID = "78ef4e7c-fe51-4573-b827-3ce3baea02ea"
TARGET_LEDGER_RUBRIC_ID = "rubric_1775876463672"
OLD_AMOUNT = "$2,000.00"
NEW_AMOUNT = "$1,900.00"
LOCAL_INSTRUCTION = (
    "Process only the unread credit memo CM-38720 from Cascade IT according to the existing "
    "Meridian AP SOP. Use the existing task tools and records, and complete the required action for this item.\n"
)

B_EMAIL_VERIFIER = r'''from pathlib import Path
import json

def verify(workspace_path, external_services_path=None):
    if external_services_path is None:
        return {"pass": False, "score": 0.0, "feedback": "external_services_path missing"}
    mailbox = Path(external_services_path) / "mailbox.json"
    if not mailbox.exists():
        return {"pass": False, "score": 0.0, "feedback": "mailbox.json not found"}
    data = json.loads(mailbox.read_text())
    for email in data.get("emails", []):
        if str(email.get("folder", "")).lower() != "sent":
            continue
        if "ar@cascadeit.com" not in str(email.get("to_addr") or "").lower():
            continue
        text = (str(email.get("subject") or "") + " " + str(email.get("body_text") or "")).lower()
        if "cm-38720" in text and any(x in text for x in ("corrected credit memo", "written explanation", "clarification")):
            return {"pass": False, "score": 0.0, "feedback": "variant B incorrectly entered dispute path for CM-38720"}
    return {"pass": True, "score": 1.0, "feedback": "variant B did not send a CM-38720 dispute email"}
'''

B_LEDGER_VERIFIER = r'''from pathlib import Path
import openpyxl

def verify(workspace_path, external_services_path=None):
    ledger = Path(workspace_path) / "ap_ledger.xlsx"
    if not ledger.exists():
        return {"pass": False, "score": 0.0, "feedback": "ap_ledger.xlsx not found"}
    wb = openpyxl.load_workbook(ledger, data_only=False)
    if "Invoice Register" not in wb.sheetnames:
        return {"pass": False, "score": 0.0, "feedback": "Invoice Register not found"}
    row = None
    for values in wb["Invoice Register"].iter_rows(min_row=2, values_only=True):
        if len(values) >= 4 and str(values[3]).strip() == "INV-38720":
            row = values
            break
    if row is None:
        return {"pass": False, "score": 0.0, "feedback": "INV-38720 row not found"}
    status = str(row[9] or "").strip().upper() if len(row) > 9 else ""
    credit = row[16] if len(row) > 16 else None
    try:
        credit = 0.0 if credit in (None, "") else float(credit)
    except (TypeError, ValueError):
        return {"pass": False, "score": 0.0, "feedback": f"unreadable Credit Applied value: {credit!r}"}
    passed = status != "DISP-HOLD" and credit == 1900.0
    return {
        "pass": passed,
        "score": 1.0 if passed else 0.0,
        "feedback": f"variant B status={status!r}, Credit Applied={credit}",
    }
'''


def _decode_ascii85_flate_stream(pdf: bytes) -> bytes:
    match = re.search(rb"stream\r?\n(.*?)(?:\r?\n)?endstream", pdf, re.S)
    if not match:
        raise RuntimeError("PDF stream not found")
    encoded = match.group(1).strip()
    try:
        compressed = base64.a85decode(b"<~" + encoded, adobe=True)
        return zlib.decompress(compressed)
    except Exception as exc:
        raise RuntimeError("Unsupported PDF stream encoding") from exc


def _rebuild_xref(pdf: bytes) -> bytes:
    xref_matches = list(re.finditer(rb"(?m)^xref\n", pdf))
    xref_pos = xref_matches[-1].start() if xref_matches else -1
    trailer_pos = pdf.find(b"trailer\n", xref_pos) if xref_pos >= 0 else -1
    startxref_pos = pdf.find(b"startxref\n", trailer_pos)
    if min(xref_pos, trailer_pos, startxref_pos) < 0:
        raise RuntimeError("PDF xref/trailer not found")

    prefix = pdf[:xref_pos]
    trailer_block = pdf[trailer_pos:startxref_pos]
    size_match = re.search(rb"/Size\s+(\d+)", trailer_block)
    if not size_match:
        raise RuntimeError("PDF /Size missing")
    size = int(size_match.group(1))

    offsets = {
        int(m.group(1)): m.start()
        for m in re.finditer(rb"(?m)^(\d+) 0 obj\b", prefix)
    }
    if any(i not in offsets for i in range(1, size)):
        raise RuntimeError("PDF object table incomplete")

    lines = [b"xref\n", f"0 {size}\n".encode(), b"0000000000 65535 f \n"]
    for i in range(1, size):
        lines.append(f"{offsets[i]:010d} 00000 n \n".encode())
    xref = b"".join(lines)
    startxref = len(prefix)
    return prefix + xref + trailer_block + b"startxref\n" + str(startxref).encode() + b"\n%%EOF\n"


def _replace_pdf_text(pdf: bytes, old: bytes, new: bytes) -> bytes:
    stream_match = re.search(rb"stream\r?\n(.*?)(?:\r?\n)?endstream", pdf, re.S)
    if not stream_match:
        raise RuntimeError("PDF stream not found")

    decoded = _decode_ascii85_flate_stream(pdf)
    if old not in decoded:
        raise RuntimeError("Expected amount not found in PDF attachment")
    decoded = decoded.replace(old, new)

    encoded = base64.a85encode(zlib.compress(decoded), adobe=True)[2:]
    pdf = pdf[:stream_match.start(1)] + encoded + pdf[stream_match.end(1):]

    obj_start = pdf.rfind(b" obj", 0, stream_match.start())
    if obj_start < 0:
        obj_start = 0
    length_match = list(re.finditer(rb"/Length\s+\d+", pdf[obj_start:stream_match.start()]))
    if not length_match:
        raise RuntimeError("PDF stream /Length not found")
    lm = length_match[-1]
    abs_start = obj_start + lm.start()
    abs_end = obj_start + lm.end()
    replacement = b"/Length " + str(len(encoded)).encode()
    pdf = pdf[:abs_start] + replacement + pdf[abs_end:]
    return _rebuild_xref(pdf)


def _rewrite_rubrics(output: Path, *, mode: str, variant: str) -> None:
    rubrics_path = output / "tests" / "rubrics.json"
    if not rubrics_path.exists():
        raise RuntimeError("tests/rubrics.json not found")
    rubrics = json.loads(rubrics_path.read_text(encoding="utf-8"))
    by_id = {str(r.get("id")): r for r in rubrics}
    missing = [rid for rid in (TARGET_EMAIL_RUBRIC_ID, TARGET_LEDGER_RUBRIC_ID) if rid not in by_id]
    if missing:
        raise RuntimeError(f"expected S01 rubric ids missing: {missing}")

    if variant == "B":
        by_id[TARGET_EMAIL_RUBRIC_ID]["rubric_text"] = (
            "S01 variant B: no CM-38720 dispute email should be sent because the $1,900 credit does not exceed the $1,950 invoice."
        )
        by_id[TARGET_EMAIL_RUBRIC_ID]["verifier_code"] = B_EMAIL_VERIFIER
        by_id[TARGET_LEDGER_RUBRIC_ID]["rubric_text"] = (
            "S01 variant B: INV-38720 must not remain on DISP-HOLD and Credit Applied must be 1900."
        )
        by_id[TARGET_LEDGER_RUBRIC_ID]["verifier_code"] = B_LEDGER_VERIFIER

    if mode == "local":
        rubrics = [
            r for r in rubrics
            if str(r.get("id")) in {TARGET_EMAIL_RUBRIC_ID, TARGET_LEDGER_RUBRIC_ID}
        ]

    rubrics_path.write_text(json.dumps(rubrics, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def materialize(source_task_dir: str | Path, output_dir: str | Path, *, mode: str, variant: str) -> None:
    source = Path(source_task_dir)
    output = Path(output_dir)
    if mode not in {"local", "full"}:
        raise ValueError("mode must be 'local' or 'full'")
    if variant not in {"A", "B"}:
        raise ValueError("variant must be 'A' or 'B'")
    if not source.is_dir():
        raise FileNotFoundError(source)

    if output.exists():
        shutil.rmtree(output)
    shutil.copytree(source, output)

    inbox_path = output / "environment/initial_external_services/google_mail/inbox.json"
    inbox = json.loads(inbox_path.read_text(encoding="utf-8"))
    target = next((e for e in inbox.get("emails", []) if str(e.get("email_id")) == TARGET_EMAIL_ID), None)
    if target is None:
        raise RuntimeError(f"Target email {TARGET_EMAIL_ID} not found")

    if mode == "local":
        (output / "instruction.md").write_text(LOCAL_INSTRUCTION, encoding="utf-8")
        for email in inbox.get("emails", []):
            email["is_read"] = True
        target["is_read"] = False

    if variant == "B":
        body = target.get("body_text") or ""
        if body.count(OLD_AMOUNT) < 2:
            raise RuntimeError("Expected two $2,000.00 references in target email body")
        target["body_text"] = body.replace(OLD_AMOUNT, NEW_AMOUNT)

        attachments = target.get("attachments") or []
        pdf_attachment = next((a for a in attachments if a.get("filename") == "CM-38720.pdf"), None)
        if pdf_attachment is None:
            raise RuntimeError("CM-38720.pdf attachment not found")
        pdf = base64.b64decode(pdf_attachment["content_base64"])
        pdf = _replace_pdf_text(pdf, OLD_AMOUNT.encode(), NEW_AMOUNT.encode())
        pdf_attachment["content_base64"] = base64.b64encode(pdf).decode()

    inbox_path.write_text(json.dumps(inbox, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    _rewrite_rubrics(output, mode=mode, variant=variant)


def _main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description="Build S01 HANDBOOK local/full A/B task variants")
    parser.add_argument("source_task_dir")
    parser.add_argument("output_dir")
    parser.add_argument("--mode", choices=["local", "full"], required=True)
    parser.add_argument("--variant", choices=["A", "B"], required=True)
    args = parser.parse_args()
    materialize(args.source_task_dir, args.output_dir, mode=args.mode, variant=args.variant)
    return 0


if __name__ == "__main__":
    raise SystemExit(_main())
