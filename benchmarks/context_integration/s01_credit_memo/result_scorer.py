from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

TARGET_INVOICE = "INV-38720"
TARGET_VENDOR_EMAIL = "ar@cascadeit.com"


def _find_file(root: Path, name: str) -> Path | None:
    direct = root / name
    if direct.exists():
        return direct
    matches = list(root.rglob(name)) if root.exists() else []
    return matches[0] if matches else None


def _load_target_row(workspace_path: Path):
    ledger_path = _find_file(workspace_path, "ap_ledger.xlsx")
    if ledger_path is None:
        return None, "ap_ledger.xlsx not found"
    wb = load_workbook(ledger_path, data_only=False)
    if "Invoice Register" not in wb.sheetnames:
        return None, "Invoice Register sheet not found"
    ws = wb["Invoice Register"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 4 and str(row[3]).strip() == TARGET_INVOICE:
            return row, None
    return None, f"{TARGET_INVOICE} row not found in Invoice Register"


def _load_mailbox(external_services_path: Path):
    candidates = [
        external_services_path / "mailbox.json",
        external_services_path / "google_mail" / "inbox.json",
    ]
    mailbox_path = next((p for p in candidates if p.exists()), None)
    if mailbox_path is None:
        matches = list(external_services_path.rglob("mailbox.json")) if external_services_path.exists() else []
        mailbox_path = matches[0] if matches else None
    if mailbox_path is None:
        return None, "mailbox.json not found"
    return json.loads(mailbox_path.read_text(encoding="utf-8")), None


def _normalize(text: str) -> str:
    text = text.lower().replace(",", "")
    return re.sub(r"\s+", " ", text).strip()


def _find_dispute_email(mailbox: dict):
    for email in mailbox.get("emails", []):
        if str(email.get("folder", "")).lower() != "sent":
            continue
        if TARGET_VENDOR_EMAIL not in str(email.get("to_addr") or "").lower():
            continue
        subject = _normalize(str(email.get("subject") or ""))
        body = _normalize(str(email.get("body_text") or ""))
        combined = f"{subject} {body}"
        if "cm-38720" not in combined:
            continue
        if any(term in combined for term in ("corrected credit memo", "written explanation", "clarification")):
            return email
    return None


def _number(value):
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        try:
            return float(str(value).replace(",", "").replace("$", ""))
        except ValueError:
            return None


def score_result(workspace_path: str | Path, external_services_path: str | Path, *, variant: str) -> dict:
    if variant not in {"A", "B"}:
        raise ValueError("variant must be 'A' or 'B'")

    row, row_error = _load_target_row(Path(workspace_path))
    if row_error:
        return {"pass": False, "score": 0.0, "feedback": row_error}
    mailbox, mail_error = _load_mailbox(Path(external_services_path))
    if mail_error:
        return {"pass": False, "score": 0.0, "feedback": mail_error}

    status = str(row[9] or "").strip().upper() if len(row) > 9 else ""
    credit_applied = _number(row[16] if len(row) > 16 else None)
    dispute_email = _find_dispute_email(mailbox)

    if variant == "A":
        checks = {
            "invoice placed on DISP-HOLD": status == "DISP-HOLD",
            "credit not applied": credit_applied == 0.0,
            "clarification email sent": dispute_email is not None,
        }
        if dispute_email is not None:
            body = _normalize(str(dispute_email.get("body_text") or ""))
            checks["email names both amounts and original invoice"] = (
                "cm-38720" in body and "inv-38720" in body and "$2000.00" in body and "$1950.00" in body
            )
    else:
        checks = {
            "invoice not left on DISP-HOLD": status != "DISP-HOLD",
            "1900 credit applied": credit_applied == 1900.0,
            "no dispute email sent": dispute_email is None,
        }

    failed = [name for name, ok in checks.items() if not ok]
    score = sum(bool(v) for v in checks.values()) / len(checks)
    return {
        "pass": not failed,
        "score": round(score, 3),
        "feedback": "all target checks passed" if not failed else "failed: " + "; ".join(failed),
        "checks": checks,
    }
