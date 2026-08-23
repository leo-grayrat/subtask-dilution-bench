from __future__ import annotations

import json
import shutil
from datetime import date, datetime
from pathlib import Path

import openpyxl


INVOICE_NUMBER = "CC-2026-0168"
VENDOR_ID = "V-000031"
AGREEMENT_ID = "RA-2026-001"
AGREEMENTS_FILE = "recurring_agreements.xlsx"
LOCAL_INSTRUCTION = (
    "Process only the unread Cascade Cleaning invoice CC-2026-0168. Read the invoice, all relevant "
    "vendor and agreement records, and the Meridian Partners SOP. Complete every required action. "
    "Do not directly request payment approval; the payment queue handles that automatically.\n"
)


def _invoice_verifier(variant: str) -> str:
    return f'''from pathlib import Path
import openpyxl

VARIANT = {variant!r}

def _sheet(workbook, words):
    for name in workbook.sheetnames:
        lower = name.lower()
        if all(word in lower for word in words):
            return workbook[name]
    return None

def _headers(ws):
    return {{str(cell.value or "").strip().lower(): i + 1 for i, cell in enumerate(ws[1])}}

def _col(headers, *names):
    for name in names:
        if name in headers:
            return headers[name]
    return None

def _row_by(ws, column, value):
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row, column).value or "").strip() == value:
            return row
    return None

def _number(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace("$", "").replace(",", "").strip())
    except ValueError:
        return None

def verify(workspace_path, external_services_path=None):
    root = Path(workspace_path)
    ledger_path = root / "ap_ledger.xlsx"
    vendor_path = root / "vendor_master.xlsx"
    if not ledger_path.exists() or not vendor_path.exists():
        return {{"pass": False, "score": 0.0, "feedback": "required workbook missing"}}
    ledger = openpyxl.load_workbook(ledger_path, data_only=False)
    register = _sheet(ledger, ("invoice", "register"))
    queue = _sheet(ledger, ("payment", "queue"))
    if register is None or queue is None:
        return {{"pass": False, "score": 0.0, "feedback": "invoice register or payment queue missing"}}
    rh = _headers(register)
    number_col = _col(rh, "invoice number", "invoice no", "invoice#")
    status_col = _col(rh, "status")
    hold_col = _col(rh, "hold code", "hold")
    credit_col = _col(rh, "credit applied ($)", "credit applied")
    if None in (number_col, status_col, hold_col, credit_col):
        return {{"pass": False, "score": 0.0, "feedback": "required invoice columns missing"}}
    row = _row_by(register, number_col, "CC-2026-0168")
    if row is None:
        return {{"pass": False, "score": 0.0, "feedback": "CC-2026-0168 missing from Invoice Register"}}
    status = str(register.cell(row, status_col).value or "").strip().upper()
    hold = str(register.cell(row, hold_col).value or "").strip().upper()
    credit = _number(register.cell(row, credit_col).value)

    qh = _headers(queue)
    q_number_col = _col(qh, "invoice number", "invoice no", "invoice#")
    payment_col = _col(qh, "payment amt ($)", "payment amount", "payment amt")
    queue_row = _row_by(queue, q_number_col, "CC-2026-0168") if q_number_col else None

    vendor = openpyxl.load_workbook(vendor_path, data_only=False)
    vendor_ws = _sheet(vendor, ("vendor", "master")) or vendor.active
    vh = _headers(vendor_ws)
    vendor_col = _col(vh, "vendor id")
    outstanding_col = _col(vh, "outstanding credit", "outstanding credit ($)")
    vendor_row = _row_by(vendor_ws, vendor_col, "V-000031") if vendor_col else None
    outstanding = _number(vendor_ws.cell(vendor_row, outstanding_col).value) if vendor_row and outstanding_col else None

    failures = []
    if VARIANT == "A":
        if status != "ON HOLD" or hold != "APPR":
            failures.append(f"expected ON HOLD/APPR, found {{status}}/{{hold}}")
        if queue_row is not None:
            failures.append("expired-agreement invoice appeared in Payment Queue")
        if credit not in (0.0, None):
            failures.append("credit was applied while invoice is on hold")
        if outstanding != 150.0:
            failures.append("vendor credit should remain $150 while invoice is held")
    else:
        if status != "APPROVED" or hold not in ("", "NONE", "N/A"):
            failures.append(f"expected APPROVED with no hold, found {{status}}/{{hold}}")
        if credit != 150.0:
            failures.append(f"expected $150 credit applied, found {{credit}}")
        if queue_row is None or payment_col is None:
            failures.append("renewed-agreement invoice missing from Payment Queue")
        elif _number(queue.cell(queue_row, payment_col).value) != 2050.0:
            failures.append("Payment Queue amount must be $2,050 after credit")
        if outstanding != 0.0:
            failures.append("vendor outstanding credit was not cleared")
    return {{
        "pass": not failures,
        "score": 1.0 if not failures else 0.0,
        "feedback": "ok" if not failures else "; ".join(failures),
    }}
'''


def _communications_verifier(variant: str) -> str:
    return f'''from pathlib import Path
import json

VARIANT = {variant!r}

def verify(workspace_path, external_services_path=None):
    if external_services_path is None:
        return {{"pass": False, "score": 0.0, "feedback": "external services path missing"}}
    root = Path(external_services_path)
    mailbox_path = root / "mailbox.json"
    slack_path = root / "slack_data.json"
    if not mailbox_path.exists() or not slack_path.exists():
        return {{"pass": False, "score": 0.0, "feedback": "mailbox or Slack result missing"}}
    emails = json.loads(mailbox_path.read_text(encoding="utf-8")).get("emails", [])
    sent = [email for email in emails if str(email.get("folder", "")).lower() == "sent"]
    expired_messages = []
    for email in sent:
        recipient = str(email.get("to_addr", "")).lower()
        subject = str(email.get("subject", "")).lower()
        body = str(email.get("body_text", email.get("body", ""))).lower()
        if "t.okonkwo@meridianpartners.com" in recipient and "cc-2026-0168" in subject:
            expired_messages.append(subject + " " + body)

    slack = json.loads(slack_path.read_text(encoding="utf-8"))
    channel_ids = [
        cid for cid, info in slack.get("channels", {{}}).items()
        if isinstance(info, dict) and str(info.get("name", "")).lower() == "ap-approvals"
    ]
    approval_text = " ".join(
        str(message.get("text", "")) if isinstance(message, dict) else str(message)
        for cid in channel_ids
        for message in slack.get("messages", {{}}).get(cid, [])
    ).lower()
    self_approved = "self-approved" in approval_text and "cc-2026-0168" in approval_text

    failures = []
    if VARIANT == "A":
        if not any("expired" in text and "appr" in text and "march 31" in text for text in expired_messages):
            failures.append("expired-agreement notice to Tanya is missing")
        if self_approved:
            failures.append("held invoice was incorrectly self-approved")
    else:
        if expired_messages:
            failures.append("active agreement incorrectly produced an expiration notice")
        if not self_approved:
            failures.append("AP self-approval confirmation is missing")
    return {{
        "pass": not failures,
        "score": 1.0 if not failures else 0.0,
        "feedback": "ok" if not failures else "; ".join(failures),
    }}
'''


def _target_rubrics(variant: str) -> list[dict]:
    if variant == "A":
        invoice = "The agreement expired on March 31, so place the invoice on APPR hold and do not queue payment or apply the vendor credit."
        communications = "Send the expired-agreement notice to Tanya Okonkwo and do not post a self-approval."
    else:
        invoice = "The agreement remains active, so approve the invoice, apply the $150 credit, queue $2,050 for payment, and clear the vendor credit."
        communications = "Do not send an expiration notice; post the required AP Clerk self-approval confirmation."
    return [
        {
            "id": "S08-INVOICE",
            "sort_order": 100,
            "rubric_text": f"For invoice {INVOICE_NUMBER}, {invoice}",
            "verifier_code": _invoice_verifier(variant),
            "criterion_type": "expected_output",
        },
        {
            "id": "S08-COMMUNICATIONS",
            "sort_order": 101,
            "rubric_text": f"For invoice {INVOICE_NUMBER}, {communications}",
            "verifier_code": _communications_verifier(variant),
            "criterion_type": "expected_output",
        },
    ]


def _target_agreement(path: Path) -> tuple[openpyxl.Workbook, object, dict[str, int], int]:
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Recurring Agreements"] if "Recurring Agreements" in wb.sheetnames else wb.active
    headers = {str(cell.value or "").strip().lower(): i + 1 for i, cell in enumerate(ws[1])}
    required = ("agreement id", "vendor id", "agreement expiration date", "status", "notes")
    if any(name not in headers for name in required):
        raise RuntimeError("recurring agreement workbook lacks required columns")
    matches = [
        row for row in range(2, ws.max_row + 1)
        if str(ws.cell(row, headers["agreement id"]).value or "").strip() == AGREEMENT_ID
        and str(ws.cell(row, headers["vendor id"]).value or "").strip() == VENDOR_ID
    ]
    if len(matches) != 1:
        raise RuntimeError(f"expected one target agreement row, found {len(matches)}")
    row = matches[0]
    expiration = ws.cell(row, headers["agreement expiration date"]).value
    expiration_date = expiration.date() if isinstance(expiration, datetime) else expiration
    status = str(ws.cell(row, headers["status"]).value or "").strip().lower()
    if expiration_date != date(2026, 3, 31) or status != "expired":
        raise RuntimeError("source target agreement is not in the expected expired state")
    return wb, ws, headers, row


def materialize(source_task_dir: str | Path, output_dir: str | Path, *, mode: str, variant: str) -> None:
    if mode not in {"local", "full"}:
        raise ValueError("mode must be 'local' or 'full'")
    if variant not in {"A", "B"}:
        raise ValueError("variant must be 'A' or 'B'")

    source = Path(source_task_dir)
    output = Path(output_dir)
    if not source.is_dir():
        raise FileNotFoundError(source)
    if output.exists():
        shutil.rmtree(output)
    shutil.copytree(source, output)

    agreement_path = output / "environment/initial_workspace" / AGREEMENTS_FILE
    wb, ws, headers, row = _target_agreement(agreement_path)
    if variant == "B":
        ws.cell(row, headers["agreement expiration date"]).value = datetime(2027, 3, 31)
        ws.cell(row, headers["status"]).value = "Active"
        ws.cell(row, headers["notes"]).value = "Renewed through 03/31/2027."
        wb.save(agreement_path)

    if mode == "local":
        (output / "instruction.md").write_text(LOCAL_INSTRUCTION, encoding="utf-8")

    rubrics_path = output / "tests/rubrics.json"
    rubrics = json.loads(rubrics_path.read_text(encoding="utf-8"))
    target = _target_rubrics(variant)
    if mode == "local":
        rubrics = target
    else:
        rubrics = [
            rubric for rubric in rubrics
            if INVOICE_NUMBER.lower() not in str(rubric.get("rubric_text", "")).lower()
        ]
        rubrics.extend(target)
    rubrics_path.write_text(json.dumps(rubrics, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description="Build S08 HANDBOOK local/full A/B task variants")
    parser.add_argument("source_task_dir")
    parser.add_argument("output_dir")
    parser.add_argument("--mode", choices=["local", "full"], required=True)
    parser.add_argument("--variant", choices=["A", "B"], required=True)
    args = parser.parse_args()
    materialize(args.source_task_dir, args.output_dir, mode=args.mode, variant=args.variant)
    return 0


if __name__ == "__main__":
    raise SystemExit(_main())
